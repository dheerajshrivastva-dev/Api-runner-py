import requests
import openpyxl
from datetime import datetime
import concurrent.futures
from tqdm import tqdm
from openpyxl.styles import PatternFill, Font


proxies = {
    'http': 'http://localhost:4000',
    'https': 'http://localhost:4000',
}

# api_list = [{'name': 'catfact.ninja', 'url': 'https://catfact.ninja/fact', 'freq': 1},
#     {'name': 'api.coindesk.com', 'url': 'https://api.coindesk.com/v1/bpi/currentprice.json', 'freq': 1},
#     {'name': 'www.boredapi.com', 'url': 'https://www.boredapi.com/api/activity', 'freq': 1},
#     {'name': 'official-joke-api.appspot.com', 'url': 'https://official-joke-api.appspot.com/random_joke', 'freq': 1},
#     {'name': 'api.agify.io', 'url': 'https://api.agify.io?name=meelad', 'freq': 1},]*100

api_list = [{'name': 'localhost', 'url': 'https://127.0.0.1:3000/test', 'freq': 1}]*4000

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.append(['API Name', 'Response Code', 'Response Time (ms) SORTED SET', 'Status', 'Extra Details'])

total_requests = sum(api['freq'] for api in api_list)

global_progress = tqdm(total=total_requests, desc="Overall Progress", unit="request", colour="green")
failed_global_progress = tqdm(total=total_requests, desc="Overall Fails", unit="request", colour="red")
success_count = {'pass': 0, 'blocked_by_proxy': 0, 'blocked_by_client': 0}

def test_api(api):
    api_name = api['name']
    api_url = api['url']
    freq = api['freq']
    
    # individual_progress = tqdm(total=freq, desc=f"Testing {api_name}", leave=False, dynamic_ncols=True)

    for _ in range(freq):
        try:
            response = requests.get(api_url, proxies=proxies, verify=False)
            response.raise_for_status()
            response_code = response.status_code
            response_time = response.elapsed.total_seconds() * 1000
            status = 'Success' if response_code == 200 else 'Other'

            sheet.append([api_name, response_code, response_time, status])

            cell = sheet.cell(row=sheet.max_row, column=4)
            if status == 'Success':
                cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            # individual_progress.update(1)
            global_progress.update(1)
            success_count['pass'] += 1
        except requests.exceptions.RequestException as e:
            if isinstance(e, requests.exceptions.ProxyError):
                sheet.append([api_name, 429, 0, 'Rejected by proxy'])

                cell = sheet.cell(row=sheet.max_row, column=4)
                cell.fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                success_count['blocked_by_proxy'] += 1
            else:
                sheet.append([api_name, other, 0, 'Rejected by Client', e])

                cell = sheet.cell(row=sheet.max_row, column=4)
                cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                success_count['blocked_by_client'] += 1

            # individual_progress.update(1)
            failed_global_progress.update(1)
        # individual_progress.close()


with concurrent.futures.ThreadPoolExecutor(max_workers=1000) as executor:
    executor.map(test_api, api_list)

global_progress.close()

for key, value in success_count.items():
    sheet.cell(row=sheet.max_row + 1, column=4, value=f"{key} : {value}")
    cell = sheet.cell(row=sheet.max_row, column=4)
    cell.font = Font(bold=True)

report_filename = f"SORTEDSET_loc_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
workbook.save(report_filename)

print("Testing completed. Report saved as:", report_filename)
