import requests
import openpyxl
from datetime import datetime
import concurrent.futures
from tqdm import tqdm

proxies = {
    'http': 'http://localhost:8080',
    'https': 'http://localhost:8080',
}

api_list = [{'name': 'catfact.ninja', 'url': 'https://catfact.ninja/fact', 'freq': 1}] * 1000

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.append(['API Name', 'Response Code', 'Response Time (ms)', 'Status'])

total_requests = sum(api['freq'] for api in api_list)

global_progress = tqdm(total=total_requests, desc="Overall Progress", unit="request", colour="green")
failed_global_progress = tqdm(total=total_requests, desc="Overall Fails", unit="request", colour="red")


def test_api(api):
    api_name = api['name']
    api_url = api['url']
    freq = api['freq']
    individual_progress = tqdm(total=freq, desc=f"Testing {api_name}", leave=False, dynamic_ncols=True)

    for _ in range(freq):
        try:
            response = requests.get(api_url, proxies=proxies)
            response.raise_for_status()
            response_code = response.status_code
            response_time = response.elapsed.total_seconds() * 1000
            status = 'Success' if response_code == 200 else 'Other'

            sheet.append([api_name, response_code, response_time, status])

            cell = sheet.cell(row=sheet.max_row, column=4)
            if status == 'Success':
                cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            individual_progress.update(1)
            global_progress.update(1)
        except requests.exceptions.RequestException as e:
            if isinstance(e, requests.exceptions.ProxyError):
                sheet.append([api_name, 429, 0, 'Rejected by proxy'])

                cell = sheet.cell(row=sheet.max_row, column=4)
                cell.fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            else:
                print('***********************************************\n')
                print(e)
                print('\n***********************************************')

            individual_progress.update(1)
            failed_global_progress.update(1)
        individual_progress.close()


with concurrent.futures.ThreadPoolExecutor(max_workers=200) as executor:
    executor.map(test_api, api_list)

global_progress.close()

report_filename = f"api_test_report_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
workbook.save(report_filename)

print("Testing completed. Report saved as:", report_filename)
